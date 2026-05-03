import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from datetime import datetime


def export_items_to_excel(items, item_type='lost'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{'丢失物品' if item_type == 'lost' else '拾获物品'}"
    
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='667EEA', end_color='667EEA', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID', '标题', '类别', '地点', '时间', '状态', '联系方式', '描述', '发布时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 30
    
    status_map = {
        'open': '寻找中' if item_type == 'lost' else '待领取',
        'closed': '已找回' if item_type == 'lost' else '已归还'
    }
    
    for row, item in enumerate(items, 2):
        time_field = item.lost_time if item_type == 'lost' else item.found_time
        
        data = [
            item.id,
            item.title,
            item.category.name if item.category else '未分类',
            item.location or '未知',
            time_field.strftime('%Y-%m-%d %H:%M') if time_field else '未知',
            status_map.get(item.status, item.status),
            item.contact_info or '未提供',
            item.description or '无描述',
            item.created_at.strftime('%Y-%m-%d %H:%M')
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    column_widths = [8, 30, 15, 25, 20, 12, 25, 40, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer


def export_matches_to_excel(matches):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "匹配记录"
    
    header_font = Font(bold=True, color='FFFFFF', size=12)
    header_fill = PatternFill(start_color='4FACFE', end_color='4FACFE', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    headers = ['ID', '丢失物品', '丢失地点', '拾获物品', '拾获地点', '匹配度', '状态', '创建时间']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    ws.row_dimensions[1].height = 30
    
    status_map = {
        'pending': '待处理',
        'confirmed': '已确认',
        'rejected': '已拒绝'
    }
    
    for row, match in enumerate(matches, 2):
        data = [
            match.id,
            match.lost_item.title if match.lost_item else '已删除',
            match.lost_item.location if match.lost_item else '-',
            match.found_item.title if match.found_item else '已删除',
            match.found_item.location if match.found_item else '-',
            f"{match.similarity_score * 100:.1f}%",
            status_map.get(match.status, match.status),
            match.created_at.strftime('%Y-%m-%d %H:%M')
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
    
    column_widths = [8, 30, 20, 30, 20, 12, 12, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer
